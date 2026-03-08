import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def kmer_similarity(seq1, seq2, k=8):
    def get_kmers(seq, k):
        return set(seq[i:i+k] for i in range(len(seq) - k + 1))
    
    kmers1 = get_kmers(seq1.upper(), k)
    kmers2 = get_kmers(seq2.upper(), k)
    
    if not kmers1 or not kmers2:
        return 0.0
    
    intersection = len(kmers1 & kmers2)
    union = len(kmers1 | kmers2)
    return intersection / union

# ---- Settings ----
MIN_SIMILARITY = 0.60  # must be at least 60% similar to something
MAX_SIMILARITY = 0.99  # but not an exact duplicate
K = 8

# ---- Load sequences ----
print("Loading sequences...")
df = pd.read_excel('output_original_2385.xlsx', header=0)
df.columns = ['Header', 'Sequence', 'Length']
sequences = df.to_dict('records')
print(f"Starting with {len(sequences)} sequences")

# ---- Keep sequences that are between 60% and 99% similar to at least one other ----
kept = []
removed = 0

for i, candidate in enumerate(sequences):
    print(f"  Checking {i+1}/{len(sequences)} | Kept: {len(kept)} | Removed: {removed}", end='\r')
    
    has_similar = False
    is_duplicate = False

    for other in sequences:
        if candidate['Header'] == other['Header']:
            continue  # skip self
        sim = kmer_similarity(candidate['Sequence'], other['Sequence'], k=K)
        if sim >= MAX_SIMILARITY:
            is_duplicate = True
            break
        if sim >= MIN_SIMILARITY:
            has_similar = True

    if has_similar and not is_duplicate:
        kept.append(candidate)
    else:
        removed += 1

print(f"\nDone! Kept {len(kept)} sequences (60%-99% similar to at least one other).")

# ---- Save to Excel ----
wb = Workbook()
ws = wb.active
ws.title = 'similar_no_duplicates'

for col, h in enumerate(['Header', 'Sequence', 'Length (bp)'], 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(name='Arial', bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='2E4057')
    cell.alignment = Alignment(horizontal='center')

for i, row in enumerate(kept):
    ws.cell(row=i+2, column=1, value=row['Header']).font = Font(name='Arial', size=9)
    ws.cell(row=i+2, column=2, value=row['Sequence']).font = Font(name='Arial', size=9)
    ws.cell(row=i+2, column=3, value=row['Length']).font = Font(name='Arial', size=9)

ws.column_dimensions['A'].width = 80
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 14

output_file = 'optrA_similar_no_duplicates.xlsx'
wb.save(output_file)
print(f"Saved to '{output_file}'")