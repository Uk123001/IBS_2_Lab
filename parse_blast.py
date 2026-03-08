import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BLAST_RESULTS  = 'blast_results.txt'
ORIGINAL_EXCEL = 'output_original_2385.xlsx'

print("Loading sequences...")
df = pd.read_excel(ORIGINAL_EXCEL, header=0)
df.columns = ['Header', 'Sequence', 'Length']

seq_dict = {}
for _, row in df.iterrows():
    key = row['Header'].split()[0]
    seq_dict[key] = row

all_ids = list(seq_dict.keys())

print("Parsing BLAST results...")
blast_cols = ['qseqid', 'sseqid', 'pident', 'length', 'mismatch',
              'gapopen', 'qstart', 'qend', 'sstart', 'send', 'evalue', 'bitscore']
blast_df = pd.read_csv(BLAST_RESULTS, sep='\t', header=None, names=blast_cols)

# Remove self-hits
blast_df = blast_df[blast_df['qseqid'] != blast_df['sseqid']]

# Keep sequences with at least one hit between 60% and 99%
has_hit = set(blast_df[(blast_df['pident'] >= 60.0) & (blast_df['pident'] < 99.0)]['qseqid'])

kept = [seq_dict[id] for id in all_ids if id in has_hit and id in seq_dict]
print(f"Started with {len(all_ids)} | Kept {len(kept)} sequences (60-99% similar to at least one other)")

wb = Workbook()
ws = wb.active
ws.title = 'BLAST_filtered'

for col, h in enumerate(['Header', 'Sequence', 'Length (bp)'], 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(name='Arial', bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='2E4057')
    cell.alignment = Alignment(horizontal='center')

for i, row in enumerate(kept):
    ws.cell(row=i+2, column=1, value=row['Header']).font = Font(name='Arial', size=9)
    ws.cell(row=i+2, column=2, value=row['Sequence']).font = Font(name='Arial', size=9)
    ws.cell(row=i+2, column=3, value=row['Length']).font = Font(name='Arial', size=9)

ws.column_dimensions['A'].width = 80
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 14

output_file = 'optrA_BLAST_filtered.xlsx'
wb.save(output_file)
print(f"Saved to '{output_file}'")